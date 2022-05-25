import openpyxl
import os
import requests
import sys
#pip install --upgrade pandas 
#pip install --upgrade openpyxl


excel_path= str(sys.argv[1])
TOKEN= str(sys.argv[2])
OWNER= str(sys.argv[3])
REPO= str(sys.argv[4])
#TOKEN = os.environ.get("GITHUB_TOKEN")
#OWNER = os.environ.get("GITHUB_OWNER")
#REPO = os.environ.get("GITHUB_REPO")
#VERSION = os.environ.get("VERSION")

#TOKEN = "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"
#OWNER = "sumitraj0103"
#REPO = "dataengineering-devops"

# Function to Trigger the Workflow

def trigger_workflow(workflow_name,baseline_number,baseline_revision):

      headers = {
        "Accept": "application/vnd.github.v3+json",
        "Authorization": f"token {TOKEN}",
      }

      data = {
        "event_type": workflow_name,
        "client_payload": {
          'baselinetag': baseline_number,
          'revision_number': baseline_revision
        }
      }
      
      print("Token :", TOKEN, "Owner :", OWNER, "Repo :", REPO)

      requests.post(
        f"https://api.github.com/repos/{OWNER}/{REPO}/dispatches",
        json=data,
        headers=headers
      )

 
# Give the location of the file
#excel_path = "C:\Projects\SCBCICD\deployment-scripts\DATAX_AUTOX_BSL_0001_Deployment_Steps_v.00.00.00.xlsx"
# workbook object is created
wb_obj = openpyxl.load_workbook(excel_path)

for ws in wb_obj.worksheets:
    print(ws.title)

print(wb_obj.worksheets[1])
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
 
#Get Detail of Baseline and Revision ID
workflows_run=[]
global baseline_number
global baseline_revision
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 3)
    #print(cell_obj.value)
    if str(cell_obj.value) == "BASELINE":
      baseline_number=(sheet_obj.cell(row = i,column= 4)).value
      print("The baseline Number is",baseline_number)
    elif str(cell_obj.value) == "REVISION":
      baseline_revision=(sheet_obj.cell(row = i,column= 4)).value
      print("The baseline Revision is",baseline_revision)

# Get the List of Workflows
# 10 is Column Number
workflows_run=[]
for i in range(10, m_row + 1):
    pipeline_cell_obj = sheet_obj.cell(row = i, column = 3)
    trigger_cell_obj= sheet_obj.cell(row = i, column = 6)
    print(pipeline_cell_obj.value)
    print(trigger_cell_obj.value)
    if (str(pipeline_cell_obj.value) != " ") and (str(trigger_cell_obj.value) == "Y"):
      workflows_run.append(cell_obj.value)


for wf_name in workflows_run:
  print("The workflow Details",wf_name)
  trigger_workflow(wf_name,baseline_number,baseline_revision)

##########Trigger Workflow ###########




